# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from core.db_manager_multiuser import DBManagerMultiUser
from utils.file_handler import FileHandler
from datetime import datetime, timedelta


def render():
    st.title('到期提醒系統')
    st.markdown('試用期滿、合約到期等提醒管理')

    # 取得當前登入用戶的 user_id
    user_id = st.session_state.user_info['user_id']

    # 使用 M6 模組專屬資料庫（支援多用戶）
    db_reminders = DBManagerMultiUser('m6_reminders', user_id=user_id)
    db_employees = DBManagerMultiUser('m6_reminders', user_id=user_id)  # M6 使用同一個資料庫，包含 employees 和 reminders 表

    today = datetime.now().date()
    seven_days_later = (today + timedelta(days=7)).strftime('%Y-%m-%d')

    # 獲取所有待處理提醒用於儀表板
    all_pending_reminders = db_reminders.get_reminders_by_range('2000-01-01', '2099-12-31', status='pending')

    # ===== 儀表板區域 =====
    st.subheader('📊 提醒儀表板')

    if all_pending_reminders and len(all_pending_reminders) > 0:
        import plotly.express as px

        chart_col1, chart_col2 = st.columns(2)

        with chart_col1:
            # 狀態分布圓餅圖
            status_counts = {'已逾期': 0, '即將到期': 0, '未來提醒': 0}
            for reminder in all_pending_reminders:
                due_date = reminder.get('due_date', '')
                if due_date < today.strftime('%Y-%m-%d'):
                    status_counts['已逾期'] += 1
                elif due_date <= seven_days_later:
                    status_counts['即將到期'] += 1
                else:
                    status_counts['未來提醒'] += 1

            status_df = pd.DataFrame({
                '狀態': list(status_counts.keys()),
                '數量': list(status_counts.values())
            })

            fig = px.pie(status_df, values='數量', names='狀態',
                        color='狀態',
                        color_discrete_map={'已逾期': '#ff4b4b', '即將到期': '#ffa500', '未來提醒': '#0068c9'},
                        title='提醒狀態分布')
            fig.update_traces(textposition='inside', textinfo='percent+label+value')
            st.plotly_chart(fig, use_container_width=True)

        with chart_col2:
            # 提醒類型分布長條圖
            type_counts = {}
            for reminder in all_pending_reminders:
                rtype = reminder.get('reminder_type', '其他')
                type_counts[rtype] = type_counts.get(rtype, 0) + 1

            type_df = pd.DataFrame({
                '提醒類型': list(type_counts.keys()),
                '數量': list(type_counts.values())
            })

            fig2 = px.bar(type_df, x='提醒類型', y='數量',
                         color='數量',
                         color_continuous_scale='Blues',
                         title='提醒類型分布')
            fig2.update_layout(showlegend=False)
            st.plotly_chart(fig2, use_container_width=True)

        # 時間軸圖表
        timeline_data = []
        for reminder in all_pending_reminders:
            due_date = reminder.get('due_date', '')
            if due_date:
                timeline_data.append({
                    '到期日': due_date,
                    '工號': reminder.get('emp_id'),
                    '姓名': reminder.get('emp_name', 'N/A'),
                    '類型': reminder.get('reminder_type', '其他')
                })

        if timeline_data:
            timeline_df = pd.DataFrame(timeline_data)
            timeline_df['到期日'] = pd.to_datetime(timeline_df['到期日'])
            timeline_df = timeline_df.sort_values('到期日')

            # 按月份統計
            timeline_df['月份'] = timeline_df['到期日'].dt.to_period('M').astype(str)
            monthly_counts = timeline_df.groupby('月份').size().reset_index(name='提醒數量')

            fig3 = px.line(monthly_counts, x='月份', y='提醒數量',
                          markers=True,
                          title='每月到期提醒數量趨勢')
            fig3.update_traces(line_color='#0068c9', line_width=3)
            st.plotly_chart(fig3, use_container_width=True)
    else:
        st.info('目前無待處理提醒，無法顯示儀表板')

    st.divider()

    # ===== 檢視切換 =====
    view = st.radio('檢視', ['待處理 (已到期 + 30天內)', '未來提醒 (30天後)', '全部項目'], horizontal=True)

    if view == '待處理 (已到期 + 30天內)':
        # 顯示從過去到未來30天內的所有待處理提醒
        start_date = '2000-01-01'  # 包含所有過去的
        end_date = (today + timedelta(days=30)).strftime('%Y-%m-%d')
        items = db_reminders.get_reminders_by_range(start_date, end_date, status='pending')
    elif view == '未來提醒 (30天後)':
        # 顯示30天後的提醒
        start_date = (today + timedelta(days=31)).strftime('%Y-%m-%d')
        end_date = '2099-12-31'
        items = db_reminders.get_reminders_by_range(start_date, end_date)
    else:
        # 全部項目
        items = db_reminders.get_reminders_by_range('2000-01-01', '2099-12-31')

    col1, col2, col3 = st.columns(3)
    with col1:
        pending_count = len([i for i in items if i.get('status') == 'pending'])
        st.metric('待處理', pending_count)
    with col2:
        urgent = [i for i in items if i.get('status') == 'pending' and i.get('due_date') and i.get('due_date') <= seven_days_later]
        st.metric('7天內到期', len(urgent))
    with col3:
        overdue = [i for i in items if i.get('status') == 'pending' and i.get('due_date') and i.get('due_date') < today.strftime('%Y-%m-%d')]
        st.metric('已逾期', len(overdue), delta=f"-{len(overdue)}" if len(overdue) > 0 else None, delta_color="inverse")

    # 匯出功能
    col_header1, col_header2 = st.columns([3, 1])
    with col_header1:
        st.subheader('待處理清單')
    with col_header2:
        if items and len([i for i in items if i.get('status') == 'pending']) > 0:
            # 準備匯出資料
            export_data = []
            for item in items:
                if item.get('status') == 'pending':
                    due_date = item.get('due_date', '')
                    if due_date < today.strftime('%Y-%m-%d'):
                        status = '已逾期'
                    elif due_date <= seven_days_later:
                        status = '即將到期'
                    else:
                        status = '未來提醒'

                    export_data.append({
                        '工號': item.get('emp_id'),
                        '姓名': item.get('emp_name'),
                        '提醒類型': item.get('reminder_type'),
                        '到期日': due_date,
                        '狀態': status,
                        '備註': item.get('notes', '')
                    })

            export_df = pd.DataFrame(export_data)

            # 轉換為 Excel 並加上格線和格式
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                export_df.to_excel(writer, index=False, sheet_name='待處理清單')

                # 取得 worksheet
                worksheet = writer.sheets['待處理清單']

                # 設定標題列格式
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')

                # 設定邊框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 設定標題列樣式
                for col_num, column_title in enumerate(export_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 設定資料列格式和邊框
                for row_num in range(2, len(export_df) + 2):
                    for col_num in range(1, len(export_df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                        # 根據狀態欄位設定顏色
                        if col_num == 5:  # 狀態欄位
                            status_value = cell.value
                            if status_value == '已逾期':
                                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                                cell.font = Font(color='FF0000', bold=True)
                            elif status_value == '即將到期':
                                cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
                                cell.font = Font(color='FF8C00', bold=True)

                # 自動調整欄寬
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

            st.download_button(
                label='📥 匯出 Excel',
                data=buffer.getvalue(),
                file_name=f'待處理提醒_{datetime.now().strftime("%Y%m%d")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    if items:
        # 先排序：已逾期 > 即將到期 > 未來
        sorted_items = sorted(
            [i for i in items if i.get('status') == 'pending'],
            key=lambda x: (x.get('due_date', '9999-12-31'))
        )

        for item in sorted_items:
            due_date = item.get('due_date', '')

            # 判斷狀態
            if due_date < today.strftime('%Y-%m-%d'):
                status_color = '🔴'
                status_text = '已逾期'
            elif due_date <= seven_days_later:
                status_color = '🟡'
                status_text = '即將到期'
            else:
                status_color = '🟢'
                status_text = '未來提醒'

            with st.container():
                col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 1, 1])

                with col1:
                    st.write(f"**{item.get('emp_id')}** - {item.get('emp_name', 'N/A')}")
                with col2:
                    st.write(f"{item.get('reminder_type', 'N/A')}")
                with col3:
                    st.write(f"到期日: {due_date}")
                with col4:
                    if status_text == '已逾期':
                        st.error(f'{status_color} {status_text}')
                    elif status_text == '即將到期':
                        st.warning(f'{status_color} {status_text}')
                    else:
                        st.info(f'{status_color} {status_text}')
                with col5:
                    if st.button('完成', key=f"done_{item.get('id')}"):
                        if db_reminders.mark_reminder_completed(item['id']):
                            st.success('已標記完成')
                            st.rerun()

                # 顯示備註
                if item.get('notes'):
                    st.caption(f"📝 {item.get('notes')}")
                st.divider()
    else:
        st.info('目前無待處理項目')
    
    with st.expander('新增提醒'):
        # 取得所有員工列表
        all_employees = db_employees.search_employee('')  # 空字串會回傳所有員工

        emp_id = ''
        emp_name = ''
        hire_date = None

        if all_employees:
            emp_options = {f"{emp['emp_id']} - {emp['name']}": emp for emp in all_employees}
            selected = st.selectbox('選擇員工', [''] + list(emp_options.keys()), key='select_emp_for_reminder')

            if selected:
                emp_data = emp_options[selected]
                emp_id = emp_data['emp_id']
                emp_name = emp_data['name']
                hire_date = emp_data.get('hire_date')

                # 顯示員工資訊
                st.info(f"工號: {emp_id} | 姓名: {emp_name} | 到職日: {hire_date if hire_date else 'N/A'}")
        else:
            st.warning('目前無員工資料，請先在「員工查詢」模組匯入員工資料或使用批次匯入功能')

        reminder_type = st.selectbox('類型', ['試用期滿', '合約到期', '其他'])

        # 自動計算到期日
        calculated_due_date = None
        if reminder_type == '試用期滿' and hire_date:
            try:
                hire_dt = pd.to_datetime(hire_date)
                probation_months = st.number_input('試用期月數', value=3, min_value=1, max_value=12, key='probation_input')
                calculated_due_date = hire_dt + pd.DateOffset(months=probation_months)
                st.success(f"📅 自動計算到期日: {calculated_due_date.strftime('%Y-%m-%d')}")
            except:
                pass

        # 到期日輸入（可以手動調整）
        if calculated_due_date:
            due_date = st.date_input('到期日 (可調整)', value=calculated_due_date.date(), key='due_date_input')
        else:
            due_date = st.date_input('到期日', key='due_date_input2')

        notes = st.text_area('備註')

        if st.button('新增', key='add_reminder_btn'):
            if emp_id and emp_name and due_date:
                success = db_reminders.add_reminder(
                    emp_id,
                    emp_name,
                    reminder_type,
                    datetime.now().strftime('%Y-%m-%d'),
                    due_date.strftime('%Y-%m-%d'),
                    notes
                )
                if success:
                    st.success('新增成功！')
                    st.rerun()
                else:
                    st.error('新增失敗！')
            else:
                st.warning('請選擇員工並填寫到期日')
    
    with st.expander('批次匯入新進人員'):
        st.write('上傳新進人員名單，自動計算試用期滿日 (到職日 + 試用期月數)')

        st.markdown("""
        **必要欄位** (支援中英文):
        - `emp_id` 或 `工號` 或 `員工編號` (必填)
        - `hire_date` 或 `到職日` 或 `入職日期` (必填)

        系統會自動計算：**到職日 + 試用期月數 = 試用期滿日**
        """)

        upload = st.file_uploader('上傳檔案', type=['xlsx', 'csv'], key='batch_import')
        probation_months = st.number_input('試用期月數', value=3, min_value=1, max_value=12)

        if upload:
            try:
                df = FileHandler.load_file(upload)
                st.write('檔案預覽:')
                st.caption('💡 點擊右上角全螢幕按鈕可查看完整資料')
                st.dataframe(df, use_container_width=True)
                st.info(f'共 {len(df)} 筆資料，{len(df.columns)} 個欄位')

                # 檢查欄位名稱
                col_map = {}
                for col in df.columns:
                    col_lower = str(col).lower().replace(' ', '')
                    if 'emp_id' in col_lower or '工號' in col or '員工編號' in col:
                        col_map['emp_id'] = col
                    if 'hire' in col_lower or '到職' in col or '入職' in col:
                        col_map['hire_date'] = col

                if 'emp_id' not in col_map or 'hire_date' not in col_map:
                    st.error(f'找不到必要欄位！請確認檔案包含「工號」和「到職日」欄位')
                    st.write(f'檔案欄位: {list(df.columns)}')
                else:
                    st.success(f'✓ 已識別欄位：工號={col_map["emp_id"]}, 到職日={col_map["hire_date"]}')

                    if st.button('執行匯入', type='primary', key='import_reminders_btn'):
                        success_count = 0
                        error_count = 0
                        error_messages = []

                        progress_bar = st.progress(0)
                        status_text = st.empty()

                        total_rows = len(df)

                        for idx, row in df.iterrows():
                            progress_bar.progress((idx + 1) / total_rows)
                            status_text.text(f'處理中: {idx + 1}/{total_rows}')

                            emp_id = row.get(col_map['emp_id'])
                            hire_date = row.get(col_map['hire_date'])
                            emp_name = row.get('姓名', row.get('Name', ''))
                            dept = row.get('部門', row.get('Department', ''))

                            if emp_id and hire_date:
                                try:
                                    # 先確保員工資料存在於兩個資料庫
                                    if emp_name:
                                        # 存入員工主資料庫
                                        db_employees.add_employee(
                                            str(emp_id),
                                            str(emp_name),
                                            None,  # national_id
                                            str(dept) if dept else None,
                                            hire_date.strftime('%Y-%m-%d') if isinstance(hire_date, pd.Timestamp) else str(hire_date)
                                        )
                                        # 存入提醒資料庫的員工表
                                        db_reminders.add_employee(
                                            str(emp_id),
                                            str(emp_name),
                                            None,
                                            str(dept) if dept else None,
                                            hire_date.strftime('%Y-%m-%d') if isinstance(hire_date, pd.Timestamp) else str(hire_date)
                                        )

                                    # 計算試用期滿日
                                    if isinstance(hire_date, str):
                                        hire_date = pd.to_datetime(hire_date)

                                    due_date = hire_date + pd.DateOffset(months=probation_months)

                                    # 新增提醒
                                    if db_reminders.add_reminder(
                                        str(emp_id),
                                        str(emp_name) if emp_name else str(emp_id),
                                        '試用期滿',
                                        datetime.now().strftime('%Y-%m-%d'),
                                        due_date.strftime('%Y-%m-%d'),
                                        f'到職日: {hire_date.strftime("%Y-%m-%d")}'
                                    ):
                                        success_count += 1
                                    else:
                                        error_count += 1
                                        error_messages.append(f'工號 {emp_id}: 新增提醒失敗')
                                except Exception as e:
                                    error_count += 1
                                    error_messages.append(f'工號 {emp_id}: {str(e)}')
                                    continue
                            else:
                                error_count += 1
                                error_messages.append(f'第 {idx + 1} 列: 缺少工號或到職日')

                        progress_bar.empty()
                        status_text.empty()

                        if success_count > 0:
                            st.success(f'匯入成功！共 {success_count} 筆提醒已建立')
                            if error_count > 0:
                                st.warning(f'有 {error_count} 筆資料匯入失敗')
                                with st.expander('查看錯誤詳情'):
                                    for msg in error_messages[:10]:  # 只顯示前10個錯誤
                                        st.text(msg)
                            st.balloons()
                            st.rerun()
                        else:
                            st.error(f'匯入失敗！所有資料都無法匯入')
                            with st.expander('查看錯誤詳情'):
                                for msg in error_messages[:10]:
                                    st.text(msg)

            except Exception as e:
                st.error(f'讀取檔案失敗: {e}')

    # === 新增：資料庫管理標籤 ===
    with st.expander('🗄️ 資料庫管理'):
        st.warning('⚠️ 請謹慎操作，刪除後無法復原！')

        # 顯示資料庫內容
        all_reminders_data = db_reminders.get_all_records()

        if all_reminders_data:
            st.info(f'提醒資料庫中共有 {len(all_reminders_data)} 筆資料')

            # 顯示資料
            df_display = pd.DataFrame(all_reminders_data)
            st.dataframe(df_display, width='stretch')

            st.divider()

            # 管理選項
            col1, col2 = st.columns(2)

            with col1:
                st.subheader('清空資料庫')
                confirm = st.checkbox('我確認要清空所有提醒資料', key='confirm_clear_all_reminders')
                if confirm:
                    if st.button('🗑️ 確認清空', type='primary', key='clear_all_reminders_btn'):
                        try:
                            db_reminders.clear_all_data()
                            st.success('提醒資料庫已清空')
                            st.rerun()
                        except Exception as e:
                            st.error(f'清空失敗: {str(e)}')

            with col2:
                st.subheader('匯出資料')
                # 匯出資料庫內容
                from io import BytesIO
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_display.to_excel(writer, index=False, sheet_name='提醒資料')
                output.seek(0)

                st.download_button(
                    label='📥 匯出資料庫內容',
                    data=output,
                    file_name=f'reminders_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

            st.divider()

            # 依條件刪除
            st.subheader('依條件刪除')

            delete_option = st.radio(
                '刪除方式',
                ['依工號刪除', '依提醒ID刪除', '依狀態批量刪除']
            )

            if delete_option == '依工號刪除':
                emp_id_to_delete = st.text_input('輸入要刪除的工號', key='delete_emp_id')
                if emp_id_to_delete and st.button('刪除此工號的所有提醒', type='primary', key='delete_by_emp'):
                    try:
                        db_reminders.delete_by_emp_id(emp_id_to_delete)
                        st.success(f'已刪除工號 {emp_id_to_delete} 的所有提醒')
                        st.rerun()
                    except Exception as e:
                        st.error(f'刪除失敗: {str(e)}')

            elif delete_option == '依提醒ID刪除':
                reminder_ids = st.multiselect(
                    '選擇要刪除的提醒',
                    options=[f"ID {r['id']}: {r['emp_id']} - {r['reminder_type']} ({r['due_date']})" for r in all_reminders_data],
                    key='delete_reminder_ids'
                )

                if reminder_ids and st.button('刪除選定提醒', type='primary', key='delete_by_ids'):
                    try:
                        for reminder_str in reminder_ids:
                            reminder_id = int(reminder_str.split(':')[0].replace('ID ', ''))
                            db_reminders.delete_reminder_by_id(reminder_id)
                        st.success(f'已刪除 {len(reminder_ids)} 筆提醒')
                        st.rerun()
                    except Exception as e:
                        st.error(f'刪除失敗: {str(e)}')

            else:  # 依狀態批量刪除
                status_to_delete = st.selectbox(
                    '選擇要刪除的狀態',
                    ['completed', 'pending']
                )

                status_count = len([r for r in all_reminders_data if r.get('status') == status_to_delete])
                st.info(f'將刪除 {status_count} 筆狀態為「{status_to_delete}」的提醒')

                if st.button(f'刪除所有「{status_to_delete}」提醒', type='primary', key='delete_by_status'):
                    confirm_batch = st.checkbox(f'我確認要刪除 {status_count} 筆提醒')
                    if confirm_batch:
                        try:
                            conn = db_reminders._get_connection()
                            cursor = conn.cursor()
                            cursor.execute("DELETE FROM reminders WHERE status = ?", (status_to_delete,))
                            conn.commit()
                            conn.close()
                            st.success(f'已刪除 {status_count} 筆提醒')
                            st.rerun()
                        except Exception as e:
                            st.error(f'刪除失敗: {str(e)}')

        else:
            st.info('提醒資料庫中無資料')
